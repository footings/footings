import os
import datetime

from attr import attrs, attrib
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import NamedStyle, Font

from footings.core.to_xlsx import obj_to_xlsx, XlsxWorkbook

from .xlsx_helpers import compare_xlsx_files


def test_obj_to_xlsx(tmp_path):

    wb = Workbook()

    # test builtins
    wksht_builtins = wb.create_sheet("test-builtins")
    obj_to_xlsx("test-bool", wksht_builtins, 2, 2, None)
    obj_to_xlsx(True, wksht_builtins, 2, 3, None)
    obj_to_xlsx("test-str", wksht_builtins, 3, 2, None)
    obj_to_xlsx("string", wksht_builtins, 3, 3, None)
    obj_to_xlsx("test-int", wksht_builtins, 4, 2, None)
    obj_to_xlsx(1, wksht_builtins, 4, 3, None)
    obj_to_xlsx("test-float", wksht_builtins, 5, 2, None)
    obj_to_xlsx(1.23456789, wksht_builtins, 5, 3, None)
    obj_to_xlsx("test-date", wksht_builtins, 6, 2, None)
    obj_to_xlsx(datetime.date(2018, 12, 31), wksht_builtins, 6, 3, None)
    obj_to_xlsx("test-datetime", wksht_builtins, 7, 2, None)
    obj_to_xlsx(datetime.datetime(2018, 12, 31, 1, 1), wksht_builtins, 7, 3, None)

    # test pd.Series
    wksht_series = wb.create_sheet("test-series")
    obj_to_xlsx(pd.Series([1, 2, 3], name="numbers"), wksht_series, 2, 2, None)
    obj_to_xlsx(pd.Series(["a", "b", "c"], name="letters"), wksht_series, 8, 2, None)

    # test pd.DataFrame
    wksht_dataframe = wb.create_sheet("test-dataframe")
    df = pd.DataFrame({"numbers": [1, 2], "letters": ["a", "b"]}, None)
    obj_to_xlsx(df, wksht_dataframe, 2, 2, None)

    # test mapping
    wksht_mapping = wb.create_sheet("test-mapping")
    obj_to_xlsx({"key": "value"}, wksht_mapping, 2, 2, None)
    obj_to_xlsx({"key": [1, 2, 3]}, wksht_mapping, 5, 2, None)
    obj_to_xlsx({"key": df}, wksht_mapping, 10, 2, None)
    obj_to_xlsx({("tuple", "key"): 1}, wksht_mapping, 15, 2, None)

    # test iterable
    wksht_iterable = wb.create_sheet("test-iterable")
    obj_to_xlsx([1, 2, 3], wksht_iterable, 2, 2, None)
    obj_to_xlsx([[1, 2, 3], [1, 2, 3]], wksht_iterable, 7, 2, None)
    obj_to_xlsx([df, df], wksht_iterable, 16, 2, None)

    # test custom
    @attrs
    class CustomOutput:
        """Custom Output"""

        a: int = attrib()
        b: int = attrib()

        def to_audit_format(self):
            return {"a": self.a, "b": self.b}

    custom1 = CustomOutput(1, 2)
    custom2 = CustomOutput(2, 4)

    wksht_custom = wb.create_sheet("test-custom")
    obj_to_xlsx(custom1, wksht_custom, 2, 2, None)
    obj_to_xlsx(custom2, wksht_custom, 6, 2, None)

    # test callable
    def test_func(a, b):
        return a, b

    wksht_callable = wb.create_sheet("test-function")
    obj_to_xlsx(test_func, wksht_callable, 2, 2, None)

    del wb["Sheet"]
    test_wb = os.path.join(tmp_path, "test-obj-to-xlsx.xlsx")
    wb.save(test_wb)

    expected_wb = os.path.join("tests", "core", "data", "expected-obj-to-xlsx.xlsx")

    assert compare_xlsx_files(test_wb, expected_wb, [], {})


def test_xlsx_workbook(tmp_path):

    bold = NamedStyle(name="bold")
    bold.font = Font(bold=True)
    wb = XlsxWorkbook.create()
    wb.create_sheet("worksheet-1", start_row=2, start_col=2)
    wb.add_named_style("bold", bold)
    wb.write_obj("worksheet-1", "B2", style="bold")
    wb.write_obj("worksheet-1", "C3", add_cols=1)
    wb.write_obj("worksheet-1", "B5", add_rows=1, style="bold")
    wb.write_obj("worksheet-1", "C7", add_rows=1, add_cols=1)

    test_wb = os.path.join(tmp_path, "test-xlsx-workbook.xlsx")
    wb.save(test_wb)

    expected_wb = os.path.join("tests", "core", "data", "expected-xlsx-workbook.xlsx")

    assert compare_xlsx_files(test_wb, expected_wb, [], {})
