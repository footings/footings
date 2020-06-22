"""Objects to turn out a model audit to an xlsx file."""

import datetime
from collections.abc import Mapping, Iterable

from attr import attrs, attrib
import pandas as pd
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import NamedStyle

#########################################################################################
# xlsx_dispatch
#########################################################################################

_ROW_SPACING = {
    "builtins": 1,
    "series": 2,
    "dataframe": 2,
    "mapping": 2,
    "iterable": 2,
}

_COL_SPACING = {
    "builtins": 0,
    "series": 0,
    "dataframe": 0,
    "mapping": 0,
    "iterable": 0,
}

_PANDAS_FORMATTING = {
    "ALIGNMENT": {"object": "", "int64": "",},
    "HEADER": {},
}


@attrs(slots=True, frozen=True, repr=False)
class XlsxRange:
    """XlsxRange"""

    row_start: int = attrib()
    col_start: int = attrib()
    row_end: int = attrib()
    col_end: int = attrib()
    row_spacing: int = attrib()
    col_spacing: int = attrib()


def _obj_to_xlsx_builtins(obj, worksheet, start_row, start_col, style):
    cell = worksheet.cell(row=start_row, column=start_col)
    cell.value = obj
    if style is not None:
        cell.style = style
    return XlsxRange(
        row_start=start_row,
        col_start=start_col,
        row_end=start_row,
        col_end=start_col,
        row_spacing=_ROW_SPACING["builtins"],
        col_spacing=_COL_SPACING["builtins"],
    )


def _obj_to_xlsx_series(obj, worksheet, start_row, start_col, style):
    _obj_to_xlsx_builtins(obj.name, worksheet, start_row, start_col, style)
    for idx, row in enumerate(obj, 1):
        _obj_to_xlsx_builtins(row, worksheet, start_row + idx, start_col, style)
    return XlsxRange(
        row_start=start_row,
        col_start=start_col,
        row_end=start_row + obj.shape[0],
        col_end=start_col,
        row_spacing=_ROW_SPACING["series"],
        col_spacing=_COL_SPACING["series"],
    )


def _obj_to_xlsx_dataframe(obj, worksheet, start_row, start_col, style):
    for idx, col in enumerate(obj):
        _obj_to_xlsx_series(obj[col], worksheet, start_row, start_col + idx, style)
    return XlsxRange(
        row_start=start_row,
        col_start=start_col,
        row_end=start_row + obj.shape[0],
        col_end=start_col,
        row_spacing=_ROW_SPACING["dataframe"],
        col_spacing=_COL_SPACING["dataframe"],
    )


def _obj_to_xlsx_mapping(obj, worksheet, start_row, start_col, style):
    if len(obj) > 0:
        for idx, (k, v) in enumerate(obj.items()):
            if not isinstance(k, str):
                try:
                    k = str(k)
                except:
                    raise ValueError("Converting key of mapping to string failed.")
            if idx == 0:
                obj_to_xlsx(k, worksheet, start_row, start_col, style)
                ret_xlsx = obj_to_xlsx(v, worksheet, start_row, start_col + 1, style)
            else:
                row = ret_xlsx.row_end + ret_xlsx.row_spacing
                obj_to_xlsx(k, worksheet, row, start_col, style)
                ret_xlsx = obj_to_xlsx(v, worksheet, row, start_col + 1, style)
        return XlsxRange(
            row_start=start_row,
            col_start=start_col,
            row_end=ret_xlsx.row_end,
            col_end=start_col,
            row_spacing=_ROW_SPACING["mapping"],
            col_spacing=_COL_SPACING["mapping"],
        )
    return XlsxRange(
        row_start=start_row,
        col_start=start_col,
        row_end=start_row,
        col_end=start_col,
        row_spacing=_ROW_SPACING["mapping"],
        col_spacing=_COL_SPACING["mapping"],
    )


def _obj_to_xlsx_iterable(obj, worksheet, start_row, start_col, style):
    if len(obj) > 0:
        for idx, x in enumerate(obj):
            if idx == 0:
                ret_xlsx = obj_to_xlsx(x, worksheet, start_row, start_col, style)
            else:
                ret_xlsx = obj_to_xlsx(
                    x,
                    worksheet,
                    ret_xlsx.row_end + ret_xlsx.row_spacing,
                    start_col,
                    style,
                )
        return XlsxRange(
            row_start=start_row,
            col_start=start_col,
            row_end=ret_xlsx.row_end,
            col_end=ret_xlsx.col_end,
            row_spacing=_ROW_SPACING["iterable"],
            col_spacing=_COL_SPACING["iterable"],
        )
    return XlsxRange(
        row_start=start_row,
        col_start=start_col,
        row_end=start_row,
        col_end=start_col,
        row_spacing=_ROW_SPACING["iterable"],
        col_spacing=_COL_SPACING["iterable"],
    )


def obj_to_xlsx(obj, worksheet, start_row, start_col, style):
    """Object to xlsx"""
    builtins = [int, float, str, datetime.date, datetime.datetime]
    if isinstance(obj, bool):
        ret = obj_to_xlsx(str(obj), worksheet, start_row, start_col, style)
    elif isinstance(obj, str) and "\n" in obj:
        new_obj = obj.split("\n")
        ret = obj_to_xlsx(new_obj, worksheet, start_row, start_col, style)
    elif any([isinstance(obj, x) for x in builtins]):
        ret = _obj_to_xlsx_builtins(obj, worksheet, start_row, start_col, style)
    elif isinstance(obj, pd.Series):
        ret = _obj_to_xlsx_series(obj, worksheet, start_row, start_col, style)
    elif isinstance(obj, pd.DataFrame):
        ret = _obj_to_xlsx_dataframe(obj, worksheet, start_row, start_col, style)
    elif isinstance(obj, Mapping):
        ret = _obj_to_xlsx_mapping(obj, worksheet, start_row, start_col, style)
    elif isinstance(obj, Iterable):
        ret = _obj_to_xlsx_iterable(obj, worksheet, start_row, start_col, style)
    elif hasattr(obj, "to_audit_format"):
        new_obj = obj.to_audit_format()
        ret = obj_to_xlsx(new_obj, worksheet, start_row, start_col, style)
    elif callable(obj):
        new_obj = "callable: " + obj.__module__ + "." + obj.__qualname__
        ret = obj_to_xlsx(new_obj, worksheet, start_row, start_col, style)
    else:
        msg = "Not able to output object to an xlsx worksheet."
        raise TypeError(msg)
    return ret


#########################################################################################
# to_excel
#########################################################################################


@attrs(slots=True, repr=False)
class XlsxWorksheet:
    """XlsxWorksheet"""

    obj: Worksheet = attrib()
    write_row: int = attrib()
    write_col: int = attrib()


@attrs(slots=True, repr=False)
class XlsxWorkbook:
    """XlsxWorkbook"""

    workbook: Workbook = attrib()
    worksheets: dict = attrib(factory=dict)
    styles: dict = attrib(factory=dict)

    @classmethod
    def create(cls):
        """Create xlsx workbook"""
        return cls(workbook=Workbook())

    def create_sheet(self, name, start_row, start_col, **kwargs):
        """Add worksheet"""
        wrksht = XlsxWorksheet(
            obj=self.workbook.create_sheet(name, **kwargs),
            write_row=start_row,
            write_col=start_col,
        )
        self.worksheets.update({name: wrksht})

    def add_named_style(self, name, style):
        """Add style"""
        if isinstance(style, NamedStyle) is False:
            msg = f"The value passsed to style is not an instance of {NamedStyle}."
            raise TypeError(msg)
        self.styles.update({name: style})

    def write_obj(self, worksheet, obj, add_rows=0, add_cols=0, style=None):
        """Write object to worksheet"""
        wrksht = self.worksheets[worksheet]
        start_row = wrksht.write_row
        start_col = wrksht.write_col
        if style is not None:
            try:
                style = self.styles.get(style)
            except KeyError as err:
                raise err
        ret_range = obj_to_xlsx(
            obj, wrksht.obj, start_row + add_rows, start_col + add_cols, style
        )
        wrksht.write_row = ret_range.row_end + ret_range.row_spacing
        wrksht.write_col = start_col + ret_range.col_spacing

    def save(self, file):
        """Close workbook"""
        del self.workbook["Sheet"]
        self.workbook.save(file)
