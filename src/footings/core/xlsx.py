import datetime
from collections.abc import Mapping, Iterable
from typing import Any, List

from attr import attrs, attrib, asdict
import pandas as pd
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import NamedStyle


@attrs(slots=True, frozen=True)
class FootingsXlsxEntry:
    """FootingsXlsxEntry"""

    worksheet: str = attrib()
    source: str = attrib()
    mapping: str = attrib()
    end_point: str = attrib()
    column_name: str = attrib()
    dtype: str = attrib()
    stable: str = attrib()
    row_start: int = attrib()
    col_start: int = attrib()
    row_end: int = attrib()
    col_end: int = attrib()


def _obj_to_xlsx_cell(obj, worksheet, row, col, **kwargs):
    if isinstance(obj, pd.Timestamp):
        if all([x == 0 for x in [obj.hour, obj.minute, obj.second]]):
            obj = str(obj.date())
        else:
            obj = str(obj)
    cell = worksheet.cell(row=row, column=col)
    cell.value = obj
    if "style" in kwargs:
        cell.style = kwargs.get("style")
    if "hyperlink" in kwargs:
        cell.hyperlink = f"#{kwargs.get('hyperlink')}"


def _obj_to_xlsx_builtins(obj, worksheet, **kwargs):
    row = worksheet.row
    col = worksheet.col
    _obj_to_xlsx_cell(obj, worksheet.obj, row, col, **kwargs)
    return [
        FootingsXlsxEntry(
            worksheet=worksheet.obj.title,
            source=kwargs.get("source", None),
            mapping=kwargs.get("mapping", None),
            end_point=kwargs.get("end_point", None),
            column_name=kwargs.get("column_name", None),
            dtype=str(type(obj)),
            stable=kwargs.get("stable", None),
            row_start=row,
            col_start=col,
            row_end=row,
            col_end=col,
        )
    ]


def _obj_to_xlsx_series(obj, worksheet, **kwargs):
    row = worksheet.row
    col = worksheet.col
    _obj_to_xlsx_cell(obj.name, worksheet.obj, row, col, **kwargs)
    for idx, row_val in enumerate(obj, 1):
        _obj_to_xlsx_cell(row_val, worksheet.obj, row + idx, col, **kwargs)
    return [
        FootingsXlsxEntry(
            worksheet=worksheet.obj.title,
            source=kwargs.get("source", None),
            mapping=kwargs.get("mapping", None),
            end_point=kwargs.get("end_point", None),
            column_name=obj.name,
            dtype=str(pd.Series),
            stable=kwargs.get("stable", None),
            row_start=row,
            col_start=col,
            row_end=row + obj.shape[0],
            col_end=col,
        )
    ]


def _obj_to_xlsx_dataframe(obj, worksheet, **kwargs):
    ret = []
    start_col = worksheet.col
    for col_nm in obj:
        ret.extend(obj_to_xlsx(obj[col_nm], worksheet, **kwargs))
        worksheet.col += 1
    worksheet.col = start_col

    return ret


def _obj_to_xlsx_mapping(obj, worksheet, **kwargs):
    ret = []
    for k, v in obj.items():
        if not isinstance(k, str):
            try:
                k = str(k)
            except:
                raise ValueError("Converting key of mapping to string failed.")
        mapping = kwargs.pop("mapping", "")
        mapping += f"[{k}]"
        kwargs.pop("end_point", None)
        key_entries = obj_to_xlsx(
            k, worksheet, mapping=mapping, end_point="KEY", **kwargs
        )
        ret.extend(key_entries)
        worksheet.col += 1
        val_entries = obj_to_xlsx(
            v, worksheet, mapping=mapping, end_point="VALUE", **kwargs
        )
        ret.extend(val_entries)
        worksheet.row = max([entry.row_end for entry in val_entries]) + 1
        worksheet.col -= 1

    return ret


def _obj_to_xlsx_iterable(obj, worksheet, **kwargs):
    ret = []
    obj_len = len(obj)
    is_nested = True if isinstance(obj[0], Iterable) else False
    for idx, x in enumerate(obj):
        if idx == obj_len - 1 and is_nested:
            worksheet.row += 1
        entries = obj_to_xlsx(x, worksheet, **kwargs)
        ret.extend(entries)
        worksheet.row = max([entry.row_end for entry in entries]) + 1

    return ret


def obj_to_xlsx(obj, worksheet, **kwargs):
    """Object to xlsx"""
    builtins = [int, float, str, datetime.date, datetime.datetime, pd.Timestamp]
    if isinstance(obj, bool):
        ret = obj_to_xlsx(str(obj), worksheet, **kwargs)
    elif isinstance(obj, str) and "\n" in obj:
        new_obj = obj.split("\n")
        ret = obj_to_xlsx(new_obj, worksheet, **kwargs)
    elif any([isinstance(obj, x) for x in builtins]):
        ret = _obj_to_xlsx_builtins(obj, worksheet, **kwargs)
    elif isinstance(obj, pd.Series):
        ret = _obj_to_xlsx_series(obj, worksheet, **kwargs)
    elif isinstance(obj, pd.DataFrame):
        ret = _obj_to_xlsx_dataframe(obj, worksheet, **kwargs)
    elif isinstance(obj, Mapping):
        if len(obj) > 0:
            ret = _obj_to_xlsx_mapping(obj, worksheet, **kwargs)
        else:
            ret = _obj_to_xlsx_builtins("{}", worksheet, **kwargs)
    elif isinstance(obj, Iterable):
        if hasattr(obj, "__iter__") and not hasattr(obj, "__len__"):
            obj = list(obj)

        if len(obj) > 0:
            ret = _obj_to_xlsx_iterable(obj, worksheet, **kwargs)
        else:
            ret = _obj_to_xlsx_builtins("[]", worksheet, **kwargs)
    elif hasattr(obj, "to_audit_xlsx"):
        new_obj = obj.to_audit_xlsx()
        ret = obj_to_xlsx(new_obj, worksheet, **kwargs)
    elif callable(obj):
        new_obj = "callable: " + obj.__module__ + "." + obj.__qualname__
        ret = _obj_to_xlsx_builtins(new_obj, worksheet, **kwargs)
    elif hasattr(obj, "__str__"):
        ret = _obj_to_xlsx_builtins(str(obj), worksheet, **kwargs)
    else:
        msg = f"Not able to output {type(obj)} to an xlsx worksheet."
        raise TypeError(msg)
    return ret


#########################################################################################
# create excel file
#########################################################################################


@attrs(slots=True, repr=False)
class FootingsXlsxSheet:
    """FootingsXlsxSheet"""

    obj: Worksheet = attrib()
    row: int = attrib(default=0)
    col: int = attrib(default=0)

    def update(self, entries: List[dict], add_rows: int = 0, add_cols: int = 0):
        self.row = max([entry.row_end for entry in entries]) + add_rows
        self.col += add_cols


@attrs(slots=True, repr=False)
class FootingsXlsxWb:
    """XlsxWorkbook"""

    workbook: Workbook = attrib()
    worksheets: dict = attrib(factory=dict)
    entries: dict = attrib(factory=list)
    styles: dict = attrib(factory=dict)

    @classmethod
    def create(cls):
        """Create xlsx workbook"""
        return cls(workbook=Workbook())

    def create_sheet(
        self, name: str, start_row: int = 0, start_col: int = 0, hidden=False
    ):
        """Add worksheet"""
        wrksht = FootingsXlsxSheet(
            obj=self.workbook.create_sheet(name), row=start_row, col=start_col,
        )
        if hidden:
            wrksht.obj.sheet_state = "hidden"
        self.worksheets.update({name: wrksht})

    def add_named_style(self, name: str, style: NamedStyle):
        """Add style"""
        if isinstance(style, NamedStyle) is False:
            msg = f"The value passsed to style is not an instance of {NamedStyle}."
            raise TypeError(msg)
        self.styles.update({name: style})

    def write_obj(
        self, worksheet: str, obj: Any, add_rows: int = 0, add_cols: int = 0, **kwargs
    ):
        """Write object to worksheet"""
        wrksht = self.worksheets[worksheet]
        entries = obj_to_xlsx(obj, wrksht, **kwargs)
        wrksht.update(entries, add_rows, add_cols)
        record_entry = kwargs.get("record_entry", True)
        if record_entry:
            self.record_entry(worksheet, entries)

    def record_entry(self, worksheet: str, entries: List[dict]):
        """Record entry to __footings__ sheet."""
        for entry in entries:
            self.entries.extend([{**asdict(entry)}])

    def save(self, file: str):
        """Close workbook"""
        self.create_sheet("__footings__", 1, 1, hidden=True)
        df = pd.DataFrame.from_records(self.entries)
        self.write_obj("__footings__", df, record_entry=False)
        del self.workbook["Sheet"]
        self.workbook.save(file)
