import datetime
from collections.abc import Iterable, Mapping
from typing import Any, List

import pandas as pd
from attr import asdict, attrib, attrs
from openpyxl import Workbook
from openpyxl.styles import Font, NamedStyle
from openpyxl.worksheet.worksheet import Worksheet


@attrs(slots=True, frozen=True)
class FootingsXlsxEntry:
    """FootingsXlsxEntry"""

    worksheet: str = attrib()
    source: str = attrib()
    mapping: str = attrib()
    end_point: str = attrib()
    column_name: str = attrib()
    dtype: str = attrib()
    stable: str = attrib()
    row_start: int = attrib()
    col_start: int = attrib()
    row_end: int = attrib()
    col_end: int = attrib()


def _obj_to_xlsx_cell(obj, worksheet, row, col, **kwargs):
    if isinstance(obj, pd.Timestamp):
        if all([x == 0 for x in [obj.hour, obj.minute, obj.second]]):
            obj = str(obj.date())
        else:
            obj = str(obj)
    cell = worksheet.cell(row=row, column=col)
    cell.value = obj
    if "style" in kwargs:
        cell.style = kwargs.get("style")
    if "hyperlink" in kwargs:
        cell.hyperlink = f"#{kwargs.get('hyperlink')}"


def _obj_to_xlsx_builtins(obj, worksheet, **kwargs):
    row = worksheet.row
    col = worksheet.col
    _obj_to_xlsx_cell(obj, worksheet.obj, row, col, **kwargs)
    return [
        FootingsXlsxEntry(
            worksheet=worksheet.obj.title,
            source=kwargs.get("source", None),
            mapping=kwargs.get("mapping", None),
            end_point=kwargs.get("end_point", None),
            column_name=kwargs.get("column_name", None),
            dtype=str(type(obj)),
            stable=kwargs.get("stable", None),
            row_start=row,
            col_start=col,
            row_end=row,
            col_end=col,
        )
    ]


def _obj_to_xlsx_series(obj, worksheet, **kwargs):
    row = worksheet.row
    col = worksheet.col
    _obj_to_xlsx_cell(obj.name, worksheet.obj, row, col, **kwargs)
    for idx, row_val in enumerate(obj, 1):
        _obj_to_xlsx_cell(row_val, worksheet.obj, row + idx, col, **kwargs)
    return [
        FootingsXlsxEntry(
            worksheet=worksheet.obj.title,
            source=kwargs.get("source", None),
            mapping=kwargs.get("mapping", None),
            end_point=kwargs.get("end_point", None),
            column_name=obj.name,
            dtype=str(pd.Series),
            stable=kwargs.get("stable", None),
            row_start=row,
            col_start=col,
            row_end=row + obj.shape[0],
            col_end=col,
        )
    ]


def _obj_to_xlsx_dataframe(obj, worksheet, **kwargs):
    ret = []
    start_col = worksheet.col
    for col_nm in obj:
        ret.extend(obj_to_xlsx(obj[col_nm], worksheet, **kwargs))
        worksheet.col += 1
    worksheet.col = start_col

    return ret


def _obj_to_xlsx_mapping(obj, worksheet, **kwargs):
    def _make_key(mapping, key):
        if mapping is None:
            return f"/{str(key)}/"
        return f"{mapping}{str(key)}/"

    ret = []
    mapping = kwargs.pop("mapping", None)
    for k, v in obj.items():
        if not isinstance(k, str):
            try:
                k = str(k)
            except:
                raise ValueError("Converting key of mapping to string failed.")
        kwargs.pop("end_point", None)
        key = _make_key(mapping, k)
        key_entries = obj_to_xlsx(k, worksheet, mapping=key, end_point="KEY", **kwargs)
        ret.extend(key_entries)
        worksheet.col += 1
        val_entries = obj_to_xlsx(v, worksheet, mapping=key, end_point="VALUE", **kwargs)
        ret.extend(val_entries)
        worksheet.row = max([entry.row_end for entry in val_entries]) + 1
        worksheet.col -= 1

    return ret


def _obj_to_xlsx_iterable(obj, worksheet, **kwargs):
    ret = []
    obj_len = len(obj)
    is_nested = True if isinstance(obj[0], Iterable) else False
    for idx, x in enumerate(obj):
        if idx == obj_len - 1 and is_nested:
            worksheet.row += 1
        entries = obj_to_xlsx(x, worksheet, **kwargs)
        ret.extend(entries)
        worksheet.row = max([entry.row_end for entry in entries]) + 1

    return ret


def obj_to_xlsx(obj, worksheet, **kwargs):
    """Object to xlsx"""
    builtins = [int, float, str, datetime.date, datetime.datetime, pd.Timestamp]
    if isinstance(obj, bool):
        ret = obj_to_xlsx(str(obj), worksheet, **kwargs)
    elif isinstance(obj, str) and "\n" in obj:
        new_obj = obj.split("\n")
        ret = obj_to_xlsx(new_obj, worksheet, **kwargs)
    elif any([isinstance(obj, x) for x in builtins]):
        ret = _obj_to_xlsx_builtins(obj, worksheet, **kwargs)
    elif isinstance(obj, pd.Series):
        ret = _obj_to_xlsx_series(obj, worksheet, **kwargs)
    elif isinstance(obj, pd.DataFrame):
        ret = _obj_to_xlsx_dataframe(obj, worksheet, **kwargs)
    elif isinstance(obj, Mapping):
        if len(obj) > 0:
            ret = _obj_to_xlsx_mapping(obj, worksheet, **kwargs)
        else:
            ret = _obj_to_xlsx_builtins("{}", worksheet, **kwargs)
    elif isinstance(obj, Iterable):
        if hasattr(obj, "__iter__") and not hasattr(obj, "__len__"):
            obj = list(obj)

        if len(obj) > 0:
            ret = _obj_to_xlsx_iterable(obj, worksheet, **kwargs)
        else:
            ret = _obj_to_xlsx_builtins("[]", worksheet, **kwargs)
    elif hasattr(obj, "to_audit_xlsx"):
        new_obj = obj.to_audit_xlsx()
        ret = obj_to_xlsx(new_obj, worksheet, **kwargs)
    elif callable(obj):
        new_obj = "callable: " + obj.__module__ + "." + obj.__qualname__
        ret = _obj_to_xlsx_builtins(new_obj, worksheet, **kwargs)
    elif hasattr(obj, "__str__"):
        ret = _obj_to_xlsx_builtins(str(obj), worksheet, **kwargs)
    else:
        msg = f"Not able to output {type(obj)} to an xlsx worksheet."
        raise TypeError(msg)
    return ret


#########################################################################################
# create excel file
#########################################################################################


@attrs(slots=True, repr=False)
class FootingsXlsxSheet:
    """FootingsXlsxSheet"""

    obj: Worksheet = attrib()
    row: int = attrib(default=0)
    col: int = attrib(default=0)

    def update(self, entries: List[dict], add_rows: int = 0, add_cols: int = 0):
        self.row = max([entry.row_end for entry in entries]) + add_rows
        self.col += add_cols


@attrs(slots=True, repr=False)
class FootingsXlsxWb:
    """XlsxWorkbook"""

    workbook: Workbook = attrib()
    worksheets: dict = attrib(factory=dict)
    entries: dict = attrib(factory=list)
    styles: dict = attrib(factory=dict)

    @classmethod
    def create(cls):
        """Create xlsx workbook"""
        return cls(workbook=Workbook())

    def create_sheet(
        self, name: str, start_row: int = 0, start_col: int = 0, hidden=False
    ):
        """Add worksheet"""
        wrksht = FootingsXlsxSheet(
            obj=self.workbook.create_sheet(name), row=start_row, col=start_col,
        )
        if hidden:
            wrksht.obj.sheet_state = "hidden"
        self.worksheets.update({name: wrksht})

    def add_named_style(self, name: str, style: NamedStyle):
        """Add style"""
        if isinstance(style, NamedStyle) is False:
            msg = f"The value passsed to style is not an instance of {NamedStyle}."
            raise TypeError(msg)
        self.styles.update({name: style})

    def write_obj(
        self, worksheet: str, obj: Any, add_rows: int = 0, add_cols: int = 0, **kwargs
    ):
        """Write object to worksheet"""
        wrksht = self.worksheets[worksheet]
        entries = obj_to_xlsx(obj, wrksht, **kwargs)
        wrksht.update(entries, add_rows, add_cols)
        record_entry = kwargs.get("record_entry", True)
        if record_entry:
            self.record_entry(worksheet, entries)

    def record_entry(self, worksheet: str, entries: List[dict]):
        """Record entry to __footings__ sheet."""
        for entry in entries:
            self.entries.extend([{**asdict(entry)}])

    def save(self, file: str):
        """Close workbook"""
        self.create_sheet("__footings__", 1, 1, hidden=True)
        df = pd.DataFrame.from_records(self.entries)
        self.write_obj("__footings__", df, record_entry=False)
        del self.workbook["Sheet"]
        self.workbook.save(file)


#########################################################################################
# xlsx helper functions
#########################################################################################


def _add_section(wb, sheet, section_name, section_value, source):
    wb.write_obj(
        sheet, section_name, add_cols=1, style=XLSX_FORMATS["title"], source=source
    )
    wb.write_obj(sheet, section_value, add_rows=2, add_cols=-1, source=source)


def _format_docstring(docstring):
    if docstring is None:
        return ""

    def _format_line(line, indent_len):
        if line[:indent_len] == "".join([" " for x in range(0, indent_len)]):
            return line[indent_len:]
        else:
            return line

    lines = docstring.split("\n")
    if lines[0] == "":
        lines = lines[1:]
    sections = [line for line in lines if "---" in line]
    if sections != []:
        section = sections[0]
        indent_len = len(section) - len(section.lstrip(" "))
    else:
        indent_len = 0
    return "\n".join(
        [_format_line(line, indent_len) if indent_len > 0 else line for line in lines]
    )


def _format_signature(sig):
    return sig


def _format_widths(wksht, start_column):
    for col in list(wksht.columns)[start_column:]:
        max_length = 0
        for cell in col:
            if len(str(cell.value)) > max_length:
                max_length = len(str(cell.value))
        if max_length <= 3:
            adj_width = (max_length + 1) * 1.2
        else:
            adj_width = max_length + 3
        wksht.column_dimensions[col[0].column_letter].width = adj_width


def _format_sheets(wb, sheet, format_beyond_d=False):
    wksht = wb.worksheets[sheet].obj
    wksht.sheet_view.showGridLines = False
    wksht.column_dimensions["A"].width = 2.14
    wksht.column_dimensions["B"].width = 14
    wksht.column_dimensions["C"].width = 14
    if format_beyond_d:
        _format_widths(wksht, start_column=3)


def _write_sheet(wb, sheet, step):
    _add_section(wb, sheet, "Step Name:", step["name"], "NAME")
    if "method_name" in step:
        _add_section(wb, sheet, "Method Name:", step["method_name"], "MTEHOD_NAME")
    if "docstring" in step:
        _add_section(
            wb, sheet, "Docstring:", _format_docstring(step["docstring"]), "DOCSTRING",
        )
    if "uses" in step:
        _add_section(
            wb, sheet, "Uses:", str(step["uses"]).replace("'", ""), "USES",
        )
    if "impacts" in step:
        _add_section(
            wb, sheet, "Impacts:", str(step["impacts"]).replace("'", ""), "IMPACTS",
        )
    if "metadata" in step:
        _add_section(
            wb, sheet, "Metadata:", step["metadata"], "METADATA",
        )
    if "output" in step:
        _add_section(
            wb, sheet, "Output:", step["output"], "OUTPUT",
        )


#########################################################################################
# create xlsx audit file
#########################################################################################

XLSX_FORMATS = {
    "title": NamedStyle(name="bold", font=Font(name="Calibri", bold=True)),
    "underline": NamedStyle("underline", font=Font(name="Calibri", underline="single")),
    "hyperlink": NamedStyle(
        "hyperlink", font=Font(name="Calibri", italic=True, bold=True)
    ),
}


def create_footings_xlsx_file(audit_dict, file, **kwargs):
    """Create xlsx file."""
    wb = FootingsXlsxWb.create()
    for format_nm, format_val in XLSX_FORMATS.items():
        wb.add_named_style(format_nm, format_val)

    # write main
    wb.create_sheet("Main", start_row=2, start_col=2)
    _add_section(wb, "Main", "Model Name:", audit_dict["name"], "NAME")

    if "signature" in audit_dict:
        _add_section(
            wb,
            "Main",
            "Signature:",
            _format_signature(audit_dict["signature"]),
            "SIGNATURE",
        )

    if "docstring" in audit_dict:
        _add_section(
            wb,
            "Main",
            "Docstring:",
            _format_docstring(audit_dict["docstring"]),
            "DOCSTRING",
        )

    _format_sheets(wb, "Main", format_beyond_d=False)

    # write instantiation
    wb.create_sheet("Instantiation", start_row=2, start_col=2)
    _add_section(
        wb,
        "Instantiation",
        "Instantiation:",
        audit_dict["instantiation"],
        "INSTANTIATION",
    )
    _format_widths(wb.worksheets["Instantiation"].obj, start_column=1)

    # write steps
    if "steps" in audit_dict:
        for k, v in audit_dict["steps"].items():
            name = v.get("name", None)
            wb.create_sheet(name, start_row=2, start_col=2)
            _write_sheet(wb, name, v)
            _format_sheets(wb, name, format_beyond_d=True)

    # write output
    wb.create_sheet("Output", start_row=2, start_col=2)
    _add_section(
        wb, "Output", "Output:", audit_dict["output"], "OUTPUT",
    )
    _format_sheets(wb, "Output", format_beyond_d=True)

    # save file
    wb.save(file)
