import json
import pathlib
from typing import Mapping

import pandas as pd
from openpyxl import load_workbook

from ..to_xlsx import FootingsXlsxEntry
from ..utils import dispatch_function


class FootingsAuditLoadError(Exception):
    """Error raised when loading file."""


def flatten_dict(key, value):
    if isinstance(value, Mapping):
        pass
    return (key, value)


def load_footings_json_file(file: str):
    with open(file, "r") as f:
        loaded_json = json.load(f)
    ret = dict([flatten_dict(k, v) for k, v in loaded_json.items()])
    return ret


def load_footings_xlsx_file(file: str):
    def _make_key(x: pd.Series):
        return FootingsXlsxEntry(
            worksheet=x.worksheet,
            source=x.source,
            mapping=x.mapping,
            end_point=x.end_point,
            column_name=x.column_name,
            dtype=x.dtype,
            stable=x.stable,
            row_start=x.row_start,
            col_start=x.col_start,
            row_end=x.row_end,
            col_end=x.col_end,
        )

    wb = load_workbook(file)
    expected_cols = set([x for x in dir(FootingsXlsxEntry) if x[0] != "_"])

    if "__footings__" not in wb.sheetnames:
        msg = (
            "The workbook is missing the sheet __footings__ which holds key information."
        )
        raise ValueError(msg)

    ws = wb["__footings__"]
    data = ws.values
    cols = next(data)
    if set(cols) != expected_cols:
        raise FootingsAuditLoadError(
            "The __footings__tab does not contain the correct columns."
        )
    data = list(data)
    df_log = pd.DataFrame(data, columns=cols)

    values = {}
    for record in df_log.itertuples():
        ws = wb[record.worksheet]
        if "DataFrame" in str(record.dtype) or "Series" in str(record.dtype):
            records = []
            col_range = range(record.col_start, record.col_end + 1)
            col_headers = [ws.cell(record.row_start, c).value for c in col_range]
            for row in range(record.row_start + 1, record.row_end + 1):
                row_capture = {
                    header: ws.cell(row, col).value
                    for header, col in zip(col_headers, col_range)
                }
                records.append(row_capture)
            values.update({_make_key(record): pd.DataFrame.from_records(records)})
        else:
            values.update(
                {_make_key(record): ws.cell(record.row_start, record.col_start).value}
            )

    return values


def load_footings_file(file: str):
    """Load footings file.

    Currently .json and .xlsx file extensions are supported.

    Parameters
    ----------
    file : str
        The file to load.

    Returns
    -------
    dict
        A dict representing the respective file and extension. Note different file extensions
        produce different dicts.
    """
    file_ext = pathlib.Path(file).suffix
    return _load_footings_file(file_ext=file_ext, file=file)


@dispatch_function(key_parameters=("file_ext",))
def _load_footings_file(file_ext, file):
    msg = f"No registered function to load a file with extension {file_ext}."
    raise NotImplementedError(msg)


@_load_footings_file.register(file_ext=".json")
def _(file):
    return load_footings_json_file(file)


@_load_footings_file.register(file_ext=".xlsx")
def _(file: str):
    return load_footings_xlsx_file(file)
