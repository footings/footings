from functools import singledispatch

import pandas as pd
from pandas.testing import assert_frame_equal, assert_series_equal
from openpyxl import load_workbook

from ..core.xlsx import FootingsXlsxEntry


#########################################################################################
# load excel file
#########################################################################################


class FootingsTestError(Exception):
    """Error raised when testing xlsx files."""


def _make_key(x: pd.Series):
    return FootingsXlsxEntry(
        worksheet=x.worksheet,
        source=x.source,
        mapping=x.mapping,
        end_point=x.end_point,
        column_name=x.column_name,
        dtype=x.dtype,
        stable=x.stable,
        row_start=x.row_start,
        col_start=x.col_start,
        row_end=x.row_end,
        col_end=x.col_end,
    )


def load_footings_audit_xlsx(file_path: str, **kwargs):
    """Read footings audit xlsx file.

    Parameters
    ----------
    file_path : str
        The path to the xlsx file to load.
    **kwargs
        Additional key word arguments.

    Returns
    -------
    dict
        A dict where the keys are the entries into the workbook and the values are the
        entry values from excel.

    Raises
    ------
    ValueEror
        If the workbook does not contain a sheet called __footings__.
    """
    wb = load_workbook(file_path)
    expected_cols = set([x for x in dir(FootingsXlsxEntry) if x[0] != "_"])

    if "__footings__" not in wb.sheetnames:
        msg = (
            "The workbook is missing the sheet __footings__ which holds key information."
        )
        raise ValueError(msg)

    ws = wb["__footings__"]
    data = ws.values
    cols = next(data)
    if set(cols) != expected_cols:
        raise FootingsTestError(
            "The __footings__tab does not contain the correct columns."
        )
    data = list(data)
    df_log = pd.DataFrame(data, columns=cols)

    values = {}
    for record in df_log.itertuples():
        ws = wb[record.worksheet]
        if "DataFrame" in str(record.dtype) or "Series" in str(record.dtype):
            records = []
            col_range = range(record.col_start, record.col_end + 1)
            col_headers = [ws.cell(record.row_start, c).value for c in col_range]
            for row in range(record.row_start + 1, record.row_end + 1):
                row_capture = {
                    header: ws.cell(row, col).value
                    for header, col in zip(col_headers, col_range)
                }
                records.append(row_capture)
            values.update({_make_key(record): pd.DataFrame.from_records(records)})
        else:
            values.update(
                {_make_key(record): ws.cell(record.row_start, record.col_start).value}
            )

    return values


#########################################################################################
# assert xlsx equal
#########################################################################################


@singledispatch
def compare_values(val1, val2):
    """A dispatch function to compare two values."""
    result = val1 == val2
    if result:
        msg = ""
    else:
        msg = "The values are different when tested generically."
    return result, msg


@compare_values.register(pd.DataFrame)
def _(val1, val2):
    result = True
    msg = ""
    try:
        assert_frame_equal(val1, val2)
    except AssertionError:
        result = False
        msg = "The values are different when tested using pd.assert_frame_equal."
    return result, msg


@compare_values.register(pd.Series)
def _(val1, val2):
    result = True
    msg = ""
    try:
        assert_series_equal(val1, val2)
    except AssertionError:
        result = False
        msg = "The values are different when tested using pd.assert_series_equal."
    return result, msg


def exclude_record(key, exclude):
    return any([all([getattr(key, k) == v for k, v in e.items()]) for e in exclude])


def _compare_footings_xlsx_files(result: dict, expected: dict, exclude: list, **kwargs):
    test = True
    log = {}
    keys = set(list(result.keys()) + list(expected.keys()))
    for key in keys:
        temp = True
        if len(exclude) > 0:
            if exclude_record(key, exclude):
                continue
        try:
            res_val = result[key]
        except KeyError:
            msg = "The result workbook is missing the entry."
            temp = False

        try:
            exp_val = expected[key]
        except KeyError:
            msg = "The expected workbook is missing the entry."
            temp = False

        if temp:
            if type(res_val) != type(exp_val):
                temp = False
                msg = "The workbook types are different"
            else:
                temp, msg = compare_values(res_val, exp_val)

        if temp is False:
            test = False
        log.update({key: {"result": temp, "msg": msg}})

    message = "\n".join(
        [
            f"{k} : {str(v['result'])} : {v['msg']}"
            for k, v in log.items()
            if v["result"] is False
        ]
    )
    return test, message


def assert_footings_audit_xlsx_equal(
    result: str, expected: str, exclude: list = [], **kwargs
):
    """Test two audit xlsx files to determine if they are equal.

    This function is useful for unit testing models to ensure models stay true over time.

    Parameters
    ----------
    result : str
        The new workbook to test against an expected workbook.
    expected : str
        The baseline workbook to compare the result against.
    exclude : dict
        A dict of records to exclude.
    **kwargs
        Additional parameters to pass.

    Returns
    -------
    bool
        True if the workbooks are equal else raises AssertionError.

    Raises
    ------
    AssertionError
        If the two workbooks are different with a list error entries.
    """
    wb_res = load_footings_audit_xlsx(result)
    wb_exp = load_footings_audit_xlsx(expected)

    test, message = _compare_footings_xlsx_files(
        wb_res, wb_exp, exclude=exclude, **kwargs
    )
    if test is False:
        raise AssertionError(f"\n{message}")
    return True
